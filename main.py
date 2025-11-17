import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta
import os
import json
import hashlib
import webbrowser
import re
from typing import Dict, List, Optional

class ExpenseTrackerApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("نظام إدارة مصاريف المواصلات")
        self.root.geometry("1100x800")
        self.root.configure(bg='#1a1a2e')
        
        # إعدادات الملفات
        self.users_file = "users_data.json"
        self.backup_file = "users_data_backup.json"
        
        # تحميل البيانات
        self.load_users()
        
        # متغيرات العمل
        self.current_user = None
        self.expenses = []
        self.current_receipt = None
        self.filter_active = False
        
        # عرض شاشة الدخول
        self.show_login_screen()
        
        # ربط حدث الإغلاق للحفظ التلقائي
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def load_users(self):
        """تحميل بيانات المستخدمين مع معالجة الأخطاء"""
        if os.path.exists(self.users_file):
            try:
                with open(self.users_file, 'r', encoding='utf-8') as f:
                    self.users_data = json.load(f)
                # ترقية البيانات القديمة
                self.upgrade_user_data()
            except Exception as e:
                # محاولة استرجاع النسخة الاحتياطية
                if os.path.exists(self.backup_file):
                    try:
                        with open(self.backup_file, 'r', encoding='utf-8') as f:
                            self.users_data = json.load(f)
                        messagebox.showwarning("تحذير", "تم استرجاع النسخة الاحتياطية")
                    except:
                        self.users_data = {}
                else:
                    self.users_data = {}
        else:
            self.users_data = {}
    
    def upgrade_user_data(self):
        """ترقية بيانات المستخدمين القديمة"""
        for username in self.users_data:
            user = self.users_data[username]
            # إضافة حقول جديدة إذا لم تكن موجودة
            if 'expenses' not in user:
                user['expenses'] = []
            if 'payment_method' not in user:
                user['payment_method'] = 'نقدي'
            if 'company_name' not in user:
                user['company_name'] = 'غير محدد'
    
    def save_users(self):
        """حفظ بيانات المستخدمين مع نسخة احتياطية"""
        try:
            # إنشاء نسخة احتياطية
            if os.path.exists(self.users_file):
                with open(self.users_file, 'r', encoding='utf-8') as f:
                    backup_data = f.read()
                with open(self.backup_file, 'w', encoding='utf-8') as f:
                    f.write(backup_data)
            
            # حفظ البيانات الجديدة
            with open(self.users_file, 'w', encoding='utf-8') as f:
                json.dump(self.users_data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            messagebox.showerror("خطأ", f"فشل حفظ البيانات: {e}")
    
    def hash_password(self, password: str) -> str:
        """تشفير كلمة المرور"""
        return hashlib.sha256(password.encode()).hexdigest()
    
    def validate_email(self, email: str) -> bool:
        """التحقق من صحة البريد الإلكتروني"""
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email) is not None
    
    def clear_window(self):
        """مسح كل العناصر من النافذة"""
        for widget in self.root.winfo_children():
            widget.destroy()
    
    # ==================== شاشات الدخول والتسجيل ====================
    
    def show_login_screen(self):
        """عرض شاشة تسجيل الدخول"""
        self.clear_window()
        
        main_frame = tk.Frame(self.root, bg='#1a1a2e')
        main_frame.place(relx=0.5, rely=0.5, anchor='center')
        
        # شعار التطبيق
        title_frame = tk.Frame(main_frame, bg='#16213e', padx=40, pady=20)
        title_frame.pack(pady=30)
        
        tk.Label(title_frame, text="🚗", font=('Arial', 48),
                bg='#16213e', fg='#0f3460').pack()
        tk.Label(title_frame, text="نظام إدارة مصاريف المواصلات",
                font=('Arial', 20, 'bold'),
                bg='#16213e', fg='#e94560').pack(pady=5)
        tk.Label(title_frame, text="Transportation Expense Management System",
                font=('Arial', 10),
                bg='#16213e', fg='#94a3b8').pack()
        
        # إطار تسجيل الدخول
        login_frame = tk.Frame(main_frame, bg='#16213e', padx=40, pady=30)
        login_frame.pack(pady=20)
        
        tk.Label(login_frame, text="تسجيل الدخول", font=('Arial', 16, 'bold'),
                bg='#16213e', fg='#ffffff').grid(row=0, column=0, columnspan=2, pady=20)
        
        tk.Label(login_frame, text="اسم المستخدم:", font=('Arial', 11),
                bg='#16213e', fg='#cbd5e1').grid(row=1, column=0, sticky='e', padx=10, pady=10)
        self.login_username = tk.Entry(login_frame, font=('Arial', 11), width=25,
                                       bg='#0f3460', fg='#ffffff', insertbackground='#ffffff',
                                       relief='flat', bd=5)
        self.login_username.grid(row=1, column=1, padx=10, pady=10)
        self.login_username.bind('<Return>', lambda e: self.login())
        
        tk.Label(login_frame, text="كلمة المرور:", font=('Arial', 11),
                bg='#16213e', fg='#cbd5e1').grid(row=2, column=0, sticky='e', padx=10, pady=10)
        self.login_password = tk.Entry(login_frame, font=('Arial', 11), width=25, show='●',
                                       bg='#0f3460', fg='#ffffff', insertbackground='#ffffff',
                                       relief='flat', bd=5)
        self.login_password.grid(row=2, column=1, padx=10, pady=10)
        self.login_password.bind('<Return>', lambda e: self.login())
        
        btn_frame = tk.Frame(login_frame, bg='#16213e')
        btn_frame.grid(row=3, column=0, columnspan=2, pady=20)
        
        tk.Button(btn_frame, text="تسجيل الدخول", font=('Arial', 12, 'bold'),
                 bg='#e94560', fg='#ffffff', padx=30, pady=10,
                 relief='flat', cursor='hand2',
                 command=self.login).pack(side='left', padx=10)
        
        tk.Button(btn_frame, text="إنشاء حساب جديد", font=('Arial', 12),
                 bg='#0f3460', fg='#ffffff', padx=30, pady=10,
                 relief='flat', cursor='hand2',
                 command=self.show_register_screen).pack(side='left', padx=10)
        
        tk.Label(main_frame, text="© 2025 جميع الحقوق محفوظة", font=('Arial', 9),
                bg='#1a1a2e', fg='#64748b').pack(pady=20)
        
        # تركيز على حقل اسم المستخدم
        self.login_username.focus()
    
    def show_register_screen(self):
        """عرض شاشة إنشاء حساب"""
        self.clear_window()
        
        main_frame = tk.Frame(self.root, bg='#1a1a2e')
        main_frame.place(relx=0.5, rely=0.5, anchor='center')
        
        tk.Label(main_frame, text="إنشاء حساب جديد", font=('Arial', 20, 'bold'),
                bg='#1a1a2e', fg='#e94560').pack(pady=20)
        
        form_frame = tk.Frame(main_frame, bg='#16213e', padx=40, pady=30)
        form_frame.pack(pady=20)
        
        fields = [
            ("الاسم الكامل:", 'name', 'entry'),
            ("اسم المستخدم:", 'username', 'entry'),
            ("كلمة المرور:", 'password', 'password'),
            ("تأكيد كلمة المرور:", 'confirm_password', 'password'),
            ("رقم الموظف:", 'employee_id', 'entry'),
            ("اسم الشركة:", 'company_name', 'entry'),
            ("القسم:", 'department', 'entry'),
            ("البريد الإلكتروني:", 'email', 'entry'),
        ]
        
        self.register_entries = {}
        
        for i, field_data in enumerate(fields):
            label, key, field_type = field_data
            
            tk.Label(form_frame, text=label, font=('Arial', 11),
                    bg='#16213e', fg='#cbd5e1').grid(row=i, column=0, sticky='e', padx=10, pady=8)
            
            if field_type == 'password':
                entry = tk.Entry(form_frame, font=('Arial', 11), width=30, show='●',
                               bg='#0f3460', fg='#ffffff', insertbackground='#ffffff',
                               relief='flat', bd=5)
            else:
                entry = tk.Entry(form_frame, font=('Arial', 11), width=30,
                               bg='#0f3460', fg='#ffffff', insertbackground='#ffffff',
                               relief='flat', bd=5)
            
            entry.grid(row=i, column=1, padx=10, pady=8)
            self.register_entries[key] = entry
        
        btn_frame = tk.Frame(form_frame, bg='#16213e')
        btn_frame.grid(row=len(fields), column=0, columnspan=2, pady=20)
        
        tk.Button(btn_frame, text="إنشاء الحساب", font=('Arial', 12, 'bold'),
                 bg='#22c55e', fg='#ffffff', padx=30, pady=10,
                 relief='flat', cursor='hand2',
                 command=self.register).pack(side='left', padx=10)
        
        tk.Button(btn_frame, text="العودة", font=('Arial', 12),
                 bg='#64748b', fg='#ffffff', padx=30, pady=10,
                 relief='flat', cursor='hand2',
                 command=self.show_login_screen).pack(side='left', padx=10)
    
    def register(self):
        """تسجيل مستخدم جديد"""
        data = {key: entry.get().strip() for key, entry in self.register_entries.items()}
        
        # التحقق من البيانات
        if not all([data['name'], data['username'], data['password'], data['employee_id'], data['company_name']]):
            messagebox.showerror("خطأ", "الرجاء ملء جميع الحقول المطلوبة!")
            return
        
        if len(data['username']) < 3:
            messagebox.showerror("خطأ", "اسم المستخدم يجب أن يكون 3 أحرف على الأقل!")
            return
        
        if len(data['password']) < 6:
            messagebox.showerror("خطأ", "كلمة المرور يجب أن تكون 6 أحرف على الأقل!")
            return
        
        if data['password'] != data['confirm_password']:
            messagebox.showerror("خطأ", "كلمة المرور غير متطابقة!")
            return
        
        if data['username'] in self.users_data:
            messagebox.showerror("خطأ", "اسم المستخدم موجود بالفعل!")
            return
        
        if data['email'] and not self.validate_email(data['email']):
            messagebox.showerror("خطأ", "البريد الإلكتروني غير صحيح!")
            return
        
        # حفظ المستخدم
        self.users_data[data['username']] = {
            'name': data['name'],
            'password': self.hash_password(data['password']),
            'employee_id': data['employee_id'],
            'company_name': data['company_name'],
            'department': data['department'],
            'email': data['email'],
            'payment_method': 'نقدي',
            'expenses': [],
            'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        self.save_users()
        messagebox.showinfo("نجح", "تم إنشاء الحساب بنجاح!\nيمكنك الآن تسجيل الدخول.")
        self.show_login_screen()
    
    def login(self):
        """تسجيل الدخول"""
        username = self.login_username.get().strip()
        password = self.login_password.get()
        
        if not username or not password:
            messagebox.showerror("خطأ", "الرجاء إدخال اسم المستخدم وكلمة المرور!")
            return
        
        if username not in self.users_data:
            messagebox.showerror("خطأ", "اسم المستخدم غير موجود!")
            return
        
        if self.users_data[username]['password'] != self.hash_password(password):
            messagebox.showerror("خطأ", "كلمة المرور غير صحيحة!")
            return
        
        self.current_user = self.users_data[username].copy()
        self.current_user['username'] = username
        self.expenses = self.current_user.get('expenses', []).copy()
        
        messagebox.showinfo("مرحباً", f"أهلاً بك {self.current_user['name']}!")
        self.show_main_app()
    
    # ==================== الواجهة الرئيسية ====================
    
    def show_main_app(self):
        """عرض التطبيق الرئيسي"""
        self.clear_window()
        
        # شريط علوي
        top_bar = tk.Frame(self.root, bg='#16213e', height=70)
        top_bar.pack(fill='x')
        top_bar.pack_propagate(False)
        
        tk.Label(top_bar, text="نظام إدارة مصاريف المواصلات",
                font=('Arial', 16, 'bold'),
                bg='#16213e', fg='#e94560').pack(side='left', padx=20, pady=15)
        
        user_frame = tk.Frame(top_bar, bg='#16213e')
        user_frame.pack(side='right', padx=20)
        
        tk.Label(user_frame, text=f"{self.current_user['name']}",
                font=('Arial', 11),
                bg='#16213e', fg='#cbd5e1').pack(side='left', padx=10)
        
        tk.Button(user_frame, text="البروفايل", font=('Arial', 10),
                 bg='#0f3460', fg='#ffffff', padx=10, pady=5,
                 relief='flat', cursor='hand2',
                 command=self.show_profile_window).pack(side='left', padx=5)
        
        tk.Button(user_frame, text="تسجيل خروج", font=('Arial', 10),
                 bg='#e94560', fg='#ffffff', padx=15, pady=5,
                 relief='flat', cursor='hand2',
                 command=self.logout).pack(side='left', padx=5)
        
        main_container = tk.Frame(self.root, bg='#1a1a2e')
        main_container.pack(fill='both', expand=True, padx=20, pady=20)
        
        # إطار إضافة مصروف
        expense_frame = tk.LabelFrame(main_container, text="إضافة مصروف جديد",
                                     font=('Arial', 12, 'bold'),
                                     bg='#16213e', fg='#cbd5e1',
                                     padx=20, pady=15)
        expense_frame.pack(fill='x', pady=(0, 15))
        
        row1 = tk.Frame(expense_frame, bg='#16213e')
        row1.pack(fill='x', pady=5)
        
        self.create_field(row1, "التاريخ:", 'date', datetime.now().strftime("%Y-%m-%d"))
        self.create_field(row1, "من:", 'from_location', "")
        self.create_field(row1, "إلى:", 'to_location', "")
        
        row2 = tk.Frame(expense_frame, bg='#16213e')
        row2.pack(fill='x', pady=5)
        
        tk.Label(row2, text="نوع المواصلة:", font=('Arial', 10),
                bg='#16213e', fg='#cbd5e1').pack(side='left', padx=5)
        self.transport_type = ttk.Combobox(row2, font=('Arial', 10), width=13,
                                          values=['أوبر', 'كريم', 'تاكسي', 'مترو', 'أتوبيس', 'سيارة خاصة', 'أخرى'],
                                          state='readonly')
        self.transport_type.set('أوبر')
        self.transport_type.pack(side='left', padx=5)
        
        tk.Label(row2, text="وسيلة الدفع:", font=('Arial', 10),
                bg='#16213e', fg='#cbd5e1').pack(side='left', padx=5)
        self.payment_method_choice = ttk.Combobox(row2, font=('Arial', 10), width=15,
                                                 values=['نقدي', 'فيزا', 'محفظة إلكترونية', 'إنستاباي', 'أخرى'],
                                                 state='readonly')
        self.payment_method_choice.set(self.current_user.get('payment_method', 'نقدي'))
        self.payment_method_choice.pack(side='left', padx=5)
        
        self.create_field(row2, "المبلغ:", 'amount', "")
        self.create_field(row2, "ملاحظات:", 'notes', "")
        
        row3 = tk.Frame(expense_frame, bg='#16213e')
        row3.pack(fill='x', pady=10)
        
        tk.Button(row3, text="إرفاق إيصال", font=('Arial', 10, 'bold'),
                 bg='#3b82f6', fg='#ffffff', padx=15, pady=8,
                 relief='flat', cursor='hand2',
                 command=self.attach_receipt).pack(side='left', padx=5)
        
        self.receipt_label = tk.Label(row3, text="لا يوجد إيصال",
                                     font=('Arial', 9), bg='#16213e', fg='#94a3b8')
        self.receipt_label.pack(side='left', padx=10)
        
        tk.Button(row3, text="إضافة المصروف", font=('Arial', 11, 'bold'),
                 bg='#22c55e', fg='#ffffff', padx=25, pady=8,
                 relief='flat', cursor='hand2',
                 command=self.add_expense).pack(side='right', padx=5)
        
        # قائمة المصاريف
        list_frame = tk.LabelFrame(main_container, text="قائمة المصاريف",
                                  font=('Arial', 12, 'bold'),
                                  bg='#16213e', fg='#cbd5e1',
                                  padx=10, pady=10)
        list_frame.pack(fill='both', expand=True, pady=(0, 15))
        
        # شريط البحث والفلترة
        filter_frame = tk.Frame(list_frame, bg='#16213e')
        filter_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(filter_frame, text="بحث:", font=('Arial', 10),
                bg='#16213e', fg='#cbd5e1').pack(side='left', padx=5)
        self.search_entry = tk.Entry(filter_frame, font=('Arial', 10), width=20,
                                     bg='#0f3460', fg='#ffffff', insertbackground='#ffffff')
        self.search_entry.pack(side='left', padx=5)
        self.search_entry.bind('<KeyRelease>', lambda e: self.filter_expenses())
        
        tk.Button(filter_frame, text="مسح البحث", font=('Arial', 9),
                 bg='#64748b', fg='#ffffff', padx=10, pady=5,
                 relief='flat', cursor='hand2',
                 command=self.clear_filter).pack(side='left', padx=5)
        
        tk.Label(filter_frame, text="فترة:", font=('Arial', 10),
                bg='#16213e', fg='#cbd5e1').pack(side='left', padx=(20, 5))
        self.period_filter = ttk.Combobox(filter_frame, font=('Arial', 9), width=12,
                                         values=['الكل', 'اليوم', 'هذا الأسبوع', 'هذا الشهر', 'آخر 30 يوم'],
                                         state='readonly')
        self.period_filter.set('الكل')
        self.period_filter.pack(side='left', padx=5)
        self.period_filter.bind('<<ComboboxSelected>>', lambda e: self.filter_expenses())
        
        # Treeview
        style = ttk.Style()
        style.theme_use('default')
        style.configure("Treeview",
                       background="#0f3460",
                       foreground="#ffffff",
                       fieldbackground="#0f3460",
                       rowheight=30)
        style.configure("Treeview.Heading",
                       background="#16213e",
                       foreground="#e94560",
                       font=('Arial', 10, 'bold'))
        style.map('Treeview', background=[('selected', '#e94560')])
        
        columns = ('التاريخ', 'من', 'إلى', 'النوع', 'وسيلة الدفع', 'المبلغ', 'ملاحظات', 'إيصال')
        self.tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=8)
        
        for col in columns:
            self.tree.heading(col, text=col)
            if col in ('ملاحظات',):
                self.tree.column(col, width=200, anchor='center')
            elif col in ('من', 'إلى'):
                self.tree.column(col, width=130, anchor='center')
            elif col in ('وسيلة الدفع', 'إيصال'):
                self.tree.column(col, width=100, anchor='center')
            else:
                self.tree.column(col, width=90, anchor='center')
        
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        
        # شريط الأزرار والإجمالي
        bottom_frame = tk.Frame(main_container, bg='#16213e', padx=15, pady=15)
        bottom_frame.pack(fill='x')
        
        self.total_label = tk.Label(bottom_frame, text="الإجمالي: 0.00 جنيه",
                                    font=('Arial', 16, 'bold'),
                                    bg='#fbbf24', fg='#000000',
                                    padx=25, pady=12)
        self.total_label.pack(side='left', padx=10)
        
        self.count_label = tk.Label(bottom_frame, text="عدد المصاريف: 0",
                                    font=('Arial', 11),
                                    bg='#16213e', fg='#cbd5e1',
                                    padx=15, pady=8)
        self.count_label.pack(side='left', padx=10)
        
        tk.Button(bottom_frame, text="تعديل", font=('Arial', 11),
                 bg='#0ea5e9', fg='#ffffff', padx=20, pady=10,
                 relief='flat', cursor='hand2',
                 command=self.edit_selected_expense).pack(side='left', padx=5)
        
        tk.Button(bottom_frame, text="حذف", font=('Arial', 11),
                 bg='#ef4444', fg='#ffffff', padx=20, pady=10,
                 relief='flat', cursor='hand2',
                 command=self.delete_expense).pack(side='left', padx=5)
        
        tk.Button(bottom_frame, text="إنشاء تقرير Excel", font=('Arial', 12, 'bold'),
                 bg='#22c55e', fg='#ffffff', padx=30, pady=12,
                 relief='flat', cursor='hand2',
                 command=self.create_excel_report).pack(side='right', padx=10)
        
        tk.Button(bottom_frame, text="إحصائيات", font=('Arial', 11),
                 bg='#8b5cf6', fg='#ffffff', padx=20, pady=10,
                 relief='flat', cursor='hand2',
                 command=self.show_statistics).pack(side='right', padx=5)
        
        # تعبئة البيانات
        self.refresh_treeview()
        self.update_total()
    
    def create_field(self, parent, label, key, default_value):
        """إنشاء حقل إدخال"""
        tk.Label(parent, text=label, font=('Arial', 10),
                bg='#16213e', fg='#cbd5e1').pack(side='left', padx=5)
        entry = tk.Entry(parent, font=('Arial', 10), width=16,
                        bg='#0f3460', fg='#ffffff', insertbackground='#ffffff',
                        relief='flat', bd=3)
        entry.insert(0, default_value)
        entry.pack(side='left', padx=5)
        setattr(self, key, entry)
    
    def attach_receipt(self):
        """إرفاق إيصال"""
        filename = filedialog.askopenfilename(
            title="اختر صورة الإيصال",
            filetypes=[("صور", "*.png *.jpg *.jpeg *.gif *.bmp"), ("كل الملفات", "*.*")]
        )
        if filename:
            self.current_receipt = filename
            self.receipt_label.config(text=f"{os.path.basename(filename)}", fg='#22c55e')
    
    def add_expense(self):
        """إضافة مصروف جديد"""
        if not self.from_location.get().strip() or not self.to_location.get().strip() or not self.amount.get().strip():
            messagebox.showerror("خطأ", "الرجاء ملء الحقول المطلوبة!")
            return
        
        try:
            amount = float(self.amount.get())
            if amount <= 0:
                messagebox.showerror("خطأ", "المبلغ يجب أن يكون أكبر من صفر!")
                return
        except ValueError:
            messagebox.showerror("خطأ", "المبلغ يجب أن يكون رقماً!")
            return
        
        date_str = self.date.get().strip()
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
        except Exception:
            messagebox.showerror("خطأ", "التاريخ يجب أن يكون بالشكل: YYYY-MM-DD")
            return
        
        expense = {
            'date': date_str,
            'from': self.from_location.get().strip(),
            'to': self.to_location.get().strip(),
            'type': self.transport_type.get(),
            'payment_method': self.payment_method_choice.get(),
            'amount': amount,
            'notes': self.notes.get().strip(),
            'receipt': self.current_receipt,
            'added_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        self.expenses.append(expense)
        self.save_user_expenses()
        
        receipt_status = "مرفق" if self.current_receipt else "لا يوجد"
        self.tree.insert('', 'end', values=(
            expense['date'],
            expense['from'],
            expense['to'],
            expense['type'],
            expense['payment_method'],
            f"{expense['amount']:.2f}",
            expense['notes'],
            receipt_status
        ))
        
        self.update_total()
        self.clear_expense_fields()
        messagebox.showinfo("نجح", "تم إضافة المصروف بنجاح!")
    
    def clear_expense_fields(self):
        """مسح حقول إدخال المصروف"""
        self.from_location.delete(0, tk.END)
        self.to_location.delete(0, tk.END)
        self.amount.delete(0, tk.END)
        self.notes.delete(0, tk.END)
        self.current_receipt = None
        self.receipt_label.config(text="لا يوجد إيصال", fg='#94a3b8')
        self.transport_type.set('أوبر')
        self.payment_method_choice.set(self.current_user.get('payment_method', 'نقدي'))
        self.date.delete(0, tk.END)
        self.date.insert(0, datetime.now().strftime("%Y-%m-%d"))
    
    def save_user_expenses(self):
        """حفظ مصاريف المستخدم الحالي"""
        self.current_user['expenses'] = self.expenses
        self.current_user['payment_method'] = self.payment_method_choice.get()
        self.users_data[self.current_user['username']] = self.current_user
        self.save_users()
    
    def delete_expense(self):
        """حذف مصروف محدد"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("تحذير", "الرجاء اختيار مصروف لحذفه!")
            return
        
        if not messagebox.askyesno("تأكيد", "هل أنت متأكد من حذف المصروف المحدد؟"):
            return
        
        index = self.tree.index(selected[0])
        self.tree.delete(selected[0])
        
        try:
            self.expenses.pop(index)
        except Exception:
            self.rebuild_expenses_from_tree()
        
        self.save_user_expenses()
        self.update_total()
        messagebox.showinfo("نجح", "تم حذف المصروف!")
    
    def edit_selected_expense(self):
        """تعديل مصروف محدد"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("تحذير", "الرجاء اختيار مصروف لتعديله!")
            return
        
        index = self.tree.index(selected[0])
        if index < 0 or index >= len(self.expenses):
            messagebox.showerror("خطأ", "خطأ في اختيار السطر.")
            return
        
        exp = self.expenses[index]
        edit_win = tk.Toplevel(self.root)
        edit_win.title("تعديل المصروف")
        edit_win.geometry("600x420")
        edit_win.configure(bg='#16213e')
        edit_win.grab_set()
        
        tk.Label(edit_win, text="تعديل بيانات المصروف", font=('Arial', 14, 'bold'),
                bg='#16213e', fg='#e94560').pack(pady=15)
        
        form = tk.Frame(edit_win, bg='#16213e')
        form.pack(pady=10, padx=30, fill='both', expand=True)
        
        # حقول التعديل
        fields = [
            ("التاريخ (YYYY-MM-DD):", exp.get('date', ''), 'date_e'),
            ("من:", exp.get('from', ''), 'from_e'),
            ("إلى:", exp.get('to', ''), 'to_e'),
            ("المبلغ:", str(exp.get('amount', '0')), 'amount_e'),
            ("ملاحظات:", exp.get('notes', ''), 'notes_e')
        ]
        
        entries = {}
        row = 0
        for label, value, key in fields:
            tk.Label(form, text=label, font=('Arial', 10),
                    bg='#16213e', fg='#cbd5e1').grid(row=row, column=0, sticky='e', padx=5, pady=8)
            entry = tk.Entry(form, font=('Arial', 10), width=40,
                           bg='#0f3460', fg='#ffffff', insertbackground='#ffffff')
            entry.grid(row=row, column=1, padx=5, pady=8)
            entry.insert(0, value)
            entries[key] = entry
            row += 1
        
        # نوع المواصلة
        tk.Label(form, text="نوع المواصلة:", font=('Arial', 10),
                bg='#16213e', fg='#cbd5e1').grid(row=row, column=0, sticky='e', padx=5, pady=8)
        type_cb = ttk.Combobox(form, font=('Arial', 10), width=37,
                              values=['أوبر', 'كريم', 'تاكسي', 'مترو', 'أتوبيس', 'سيارة خاصة', 'أخرى'],
                              state='readonly')
        type_cb.grid(row=row, column=1, padx=5, pady=8)
        type_cb.set(exp.get('type', 'أوبر'))
        row += 1
        
        # وسيلة الدفع
        tk.Label(form, text="وسيلة الدفع:", font=('Arial', 10),
                bg='#16213e', fg='#cbd5e1').grid(row=row, column=0, sticky='e', padx=5, pady=8)
        pay_cb = ttk.Combobox(form, font=('Arial', 10), width=37,
                             values=['نقدي', 'فيزا', 'محفظة إلكترونية', 'إنستاباي', 'أخرى'],
                             state='readonly')
        pay_cb.grid(row=row, column=1, padx=5, pady=8)
        pay_cb.set(exp.get('payment_method', 'نقدي'))
        row += 1
        
        # إرفاق إيصال
        new_receipt_path = tk.StringVar(value=exp.get('receipt') or "")
        
        def choose_new_receipt():
            fn = filedialog.askopenfilename(
                title="اختر صورة إيصال",
                filetypes=[("صور", "*.png *.jpg *.jpeg *.gif *.bmp"), ("كل الملفات", "*.*")]
            )
            if fn:
                new_receipt_path.set(fn)
                lbl_receipt.config(text=os.path.basename(fn))
        
        tk.Button(form, text="تغيير/إرفاق إيصال", font=('Arial', 9),
                 bg='#3b82f6', fg='#ffffff', padx=10, pady=6,
                 relief='flat', command=choose_new_receipt).grid(row=row, column=0, sticky='e', padx=5, pady=10)
        lbl_receipt = tk.Label(form, text=os.path.basename(exp.get('receipt') or "لا يوجد"),
                              bg='#16213e', fg='#cbd5e1')
        lbl_receipt.grid(row=row, column=1, sticky='w', padx=5, pady=10)
        
        # أزرار الحفظ والإلغاء
        btn_frame = tk.Frame(edit_win, bg='#16213e')
        btn_frame.pack(pady=15)
        
        def save_edit():
            try:
                datetime.strptime(entries['date_e'].get().strip(), "%Y-%m-%d")
            except:
                messagebox.showerror("خطأ", "التاريخ غير صحيح، استخدم YYYY-MM-DD")
                return
            try:
                amt = float(entries['amount_e'].get().strip())
                if amt <= 0:
                    raise ValueError
            except:
                messagebox.showerror("خطأ", "المبلغ يجب أن يكون رقماً أكبر من صفر")
                return
            
            updated = {
                'date': entries['date_e'].get().strip(),
                'from': entries['from_e'].get().strip(),
                'to': entries['to_e'].get().strip(),
                'type': type_cb.get(),
                'payment_method': pay_cb.get(),
                'amount': amt,
                'notes': entries['notes_e'].get().strip(),
                'receipt': new_receipt_path.get() if new_receipt_path.get() else None,
                'added_at': exp.get('added_at', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                'updated_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            self.expenses[index] = updated
            self.save_user_expenses()
            self.refresh_treeview()
            self.update_total()
            messagebox.showinfo("نجح", "تم حفظ التعديلات.")
            edit_win.destroy()
        
        tk.Button(btn_frame, text="حفظ التعديلات", font=('Arial', 11, 'bold'),
                 bg='#22c55e', fg='#ffffff', padx=25, pady=8,
                 relief='flat', command=save_edit).pack(side='left', padx=10)
        
        tk.Button(btn_frame, text="إلغاء", font=('Arial', 11),
                 bg='#64748b', fg='#ffffff', padx=25, pady=8,
                 relief='flat', command=edit_win.destroy).pack(side='left', padx=10)
    
    def rebuild_expenses_from_tree(self):
        """إعادة بناء قائمة المصاريف من Treeview"""
        items = self.tree.get_children()
        new_expenses = []
        for iid in items:
            vals = self.tree.item(iid, 'values')
            try:
                amount = float(str(vals[5]).replace(',', ''))
            except:
                amount = 0.0
            new_expenses.append({
                'date': vals[0],
                'from': vals[1],
                'to': vals[2],
                'type': vals[3],
                'payment_method': vals[4],
                'amount': amount,
                'notes': vals[6],
                'receipt': None
            })
        self.expenses = new_expenses
    
    def refresh_treeview(self):
        """تحديث عرض المصاريف"""
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        expenses_to_show = self.expenses if not self.filter_active else self.filtered_expenses
        
        for expense in expenses_to_show:
            receipt_status = "مرفق" if expense.get('receipt') else "لا يوجد"
            self.tree.insert('', 'end', values=(
                expense.get('date', ''),
                expense.get('from', ''),
                expense.get('to', ''),
                expense.get('type', ''),
                expense.get('payment_method', ''),
                f"{expense.get('amount', 0):.2f}",
                expense.get('notes', ''),
                receipt_status
            ))
    
    def filter_expenses(self):
        """فلترة المصاريف حسب البحث والفترة"""
        search_text = self.search_entry.get().strip().lower()
        period = self.period_filter.get()
        
        self.filtered_expenses = []
        today = datetime.now()
        
        for exp in self.expenses:
            # فلتر البحث
            if search_text:
                searchable = f"{exp.get('from', '')} {exp.get('to', '')} {exp.get('type', '')} {exp.get('notes', '')}".lower()
                if search_text not in searchable:
                    continue
            
            # فلتر الفترة
            if period != 'الكل':
                try:
                    exp_date = datetime.strptime(exp.get('date', ''), "%Y-%m-%d")
                    if period == 'اليوم':
                        if exp_date.date() != today.date():
                            continue
                    elif period == 'هذا الأسبوع':
                        week_start = today - timedelta(days=today.weekday())
                        if exp_date < week_start:
                            continue
                    elif period == 'هذا الشهر':
                        if exp_date.month != today.month or exp_date.year != today.year:
                            continue
                    elif period == 'آخر 30 يوم':
                        if exp_date < today - timedelta(days=30):
                            continue
                except:
                    continue
            
            self.filtered_expenses.append(exp)
        
        self.filter_active = bool(search_text) or period != 'الكل'
        self.refresh_treeview()
        self.update_total()
    
    def clear_filter(self):
        """مسح الفلتر"""
        self.search_entry.delete(0, tk.END)
        self.period_filter.set('الكل')
        self.filter_active = False
        self.refresh_treeview()
        self.update_total()
    
    def update_total(self):
        """تحديث الإجمالي وعدد المصاريف"""
        expenses_to_count = self.expenses if not self.filter_active else self.filtered_expenses
        total = sum(exp.get('amount', 0) for exp in expenses_to_count)
        count = len(expenses_to_count)
        
        self.total_label.config(text=f"الإجمالي: {total:.2f} جنيه")
        self.count_label.config(text=f"عدد المصاريف: {count}")
    
    def show_statistics(self):
        """عرض نافذة الإحصائيات"""
        if not self.expenses:
            messagebox.showinfo("معلومة", "لا توجد مصاريف لعرض الإحصائيات!")
            return
        
        stats_win = tk.Toplevel(self.root)
        stats_win.title("إحصائيات المصاريف")
        stats_win.geometry("600x500")
        stats_win.configure(bg='#16213e')
        stats_win.grab_set()
        
        tk.Label(stats_win, text="إحصائيات المصاريف", font=('Arial', 16, 'bold'),
                bg='#16213e', fg='#e94560').pack(pady=15)
        
        frame = tk.Frame(stats_win, bg='#16213e')
        frame.pack(pady=10, padx=30, fill='both', expand=True)
        
        # حساب الإحصائيات
        total = sum(exp.get('amount', 0) for exp in self.expenses)
        count = len(self.expenses)
        avg = total / count if count > 0 else 0
        
        # أعلى وأقل مصروف
        amounts = [exp.get('amount', 0) for exp in self.expenses]
        max_expense = max(amounts) if amounts else 0
        min_expense = min(amounts) if amounts else 0
        
        # حسب نوع المواصلة
        by_type = {}
        for exp in self.expenses:
            t = exp.get('type', 'أخرى')
            by_type[t] = by_type.get(t, 0) + exp.get('amount', 0)
        
        # حسب وسيلة الدفع
        by_payment = {}
        for exp in self.expenses:
            p = exp.get('payment_method', 'نقدي')
            by_payment[p] = by_payment.get(p, 0) + exp.get('amount', 0)
        
        # عرض الإحصائيات
        stats = [
            ("إجمالي المصاريف:", f"{total:.2f} جنيه"),
            ("عدد المصاريف:", str(count)),
            ("متوسط المصروف:", f"{avg:.2f} جنيه"),
            ("أعلى مصروف:", f"{max_expense:.2f} جنيه"),
            ("أقل مصروف:", f"{min_expense:.2f} جنيه"),
        ]
        
        row = 0
        for label, value in stats:
            tk.Label(frame, text=label, font=('Arial', 11, 'bold'),
                    bg='#16213e', fg='#cbd5e1').grid(row=row, column=0, sticky='e', padx=10, pady=8)
            tk.Label(frame, text=value, font=('Arial', 11),
                    bg='#16213e', fg='#ffffff').grid(row=row, column=1, sticky='w', padx=10, pady=8)
            row += 1
        
        # حسب نوع المواصلة
        tk.Label(frame, text="حسب نوع المواصلة:", font=('Arial', 12, 'bold'),
                bg='#16213e', fg='#e94560').grid(row=row, column=0, columnspan=2, pady=(20, 10))
        row += 1
        
        for t_type, t_amount in sorted(by_type.items(), key=lambda x: x[1], reverse=True):
            percentage = (t_amount / total * 100) if total > 0 else 0
            tk.Label(frame, text=f"{t_type}:", font=('Arial', 10),
                    bg='#16213e', fg='#cbd5e1').grid(row=row, column=0, sticky='e', padx=10, pady=5)
            tk.Label(frame, text=f"{t_amount:.2f} جنيه ({percentage:.1f}%)", font=('Arial', 10),
                    bg='#16213e', fg='#ffffff').grid(row=row, column=1, sticky='w', padx=10, pady=5)
            row += 1
        
        # حسب وسيلة الدفع
        tk.Label(frame, text="حسب وسيلة الدفع:", font=('Arial', 12, 'bold'),
                bg='#16213e', fg='#e94560').grid(row=row, column=0, columnspan=2, pady=(20, 10))
        row += 1
        
        for p_method, p_amount in sorted(by_payment.items(), key=lambda x: x[1], reverse=True):
            percentage = (p_amount / total * 100) if total > 0 else 0
            tk.Label(frame, text=f"{p_method}:", font=('Arial', 10),
                    bg='#16213e', fg='#cbd5e1').grid(row=row, column=0, sticky='e', padx=10, pady=5)
            tk.Label(frame, text=f"{p_amount:.2f} جنيه ({percentage:.1f}%)", font=('Arial', 10),
                    bg='#16213e', fg='#ffffff').grid(row=row, column=1, sticky='w', padx=10, pady=5)
            row += 1
        
        tk.Button(stats_win, text="إغلاق", font=('Arial', 11),
                 bg='#64748b', fg='#ffffff', padx=30, pady=10,
                 relief='flat', command=stats_win.destroy).pack(pady=15)
    
    def create_excel_report(self):
        """إنشاء تقرير Excel"""
        if not self.expenses:
            messagebox.showerror("خطأ", "لا توجد مصاريف لإنشاء التقرير!")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"تقرير_مصاريف_{self.current_user['username']}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            self.generate_excel(filename)
            if messagebox.askyesno("نجح", f"تم إنشاء التقرير بنجاح!\nهل تريد فتح الملف؟"):
                webbrowser.open(f'file://{os.path.abspath(filename)}')
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء إنشاء التقرير:\n{str(e)}")
    
    def generate_excel(self, filename):
        """إنشاء ملف Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تقرير المصاريف"
        
        # تنسيق الأعمدة
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 20
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # العنوان
        ws.merge_cells('A1:H1')
        header_cell = ws['A1']
        header_cell.value = "تقرير مصاريف المواصلات والانتقالات"
        header_cell.font = Font(size=16, bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # معلومات المستخدم
        row = 3
        info_style = Font(size=11, bold=True)
        info_data = [
            ("اسم الموظف:", self.current_user['name']),
            ("رقم الموظف:", self.current_user['employee_id']),
            ("اسم الشركة:", self.current_user.get('company_name', 'غير محدد')),
            ("القسم:", self.current_user.get('department', '')),
            ("وسيلة الدفع الافتراضية:", self.current_user.get('payment_method', 'نقدي')),
            ("تاريخ التقرير:", datetime.now().strftime("%Y-%m-%d %H:%M"))
        ]
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = info_style
            row += 1
        
        # رأس الجدول
        row += 1
        headers = ['التاريخ', 'من', 'إلى', 'نوع المواصلة', 'وسيلة الدفع', 'المبلغ (جنيه)', 'ملاحظات', 'الإيصال']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(size=11, bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        ws.row_dimensions[row].height = 25
        data_start_row = row + 1
        current_row = data_start_row
        
        # بيانات المصاريف
        for expense in self.expenses:
            ws[f'A{current_row}'] = expense.get('date', '')
            ws[f'B{current_row}'] = expense.get('from', '')
            ws[f'C{current_row}'] = expense.get('to', '')
            ws[f'D{current_row}'] = expense.get('type', '')
            ws[f'E{current_row}'] = expense.get('payment_method', '')
            ws[f'F{current_row}'] = expense.get('amount', 0)
            ws[f'G{current_row}'] = expense.get('notes', '')
            
            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                
                if (current_row - data_start_row) % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # إضافة الإيصال
            receipt_path = expense.get('receipt')
            if receipt_path and os.path.exists(receipt_path):
                try:
                    img = Image(receipt_path)
                    max_dim = 150
                    if img.width > max_dim or img.height > max_dim:
                        ratio = min(max_dim / img.width, max_dim / img.height)
                        img.width = int(img.width * ratio)
                        img.height = int(img.height * ratio)
                    ws.add_image(img, f'H{current_row}')
                    ws.row_dimensions[current_row].height = max(115, int(img.height * 0.75) + 10)
                    ws[f'H{current_row}'] = "مرفق"
                except:
                    ws[f'H{current_row}'] = "خطأ في الصورة"
            else:
                ws[f'H{current_row}'] = "لا يوجد"
            
            ws[f'H{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'H{current_row}'].border = thin_border
            current_row += 1
        
        # الإجمالي
        total_row = current_row + 1
        ws.merge_cells(f'A{total_row}:E{total_row}')
        total_label = ws[f'A{total_row}']
        total_label.value = "الإجمالي الكلي"
        total_label.font = Font(size=12, bold=True)
        total_label.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        total_label.alignment = Alignment(horizontal='center', vertical='center')
        total_label.border = thin_border
        
        total_amount = sum(exp.get('amount', 0) for exp in self.expenses)
        total_cell = ws[f'F{total_row}']
        total_cell.value = total_amount
        total_cell.font = Font(size=12, bold=True)
        total_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        total_cell.alignment = Alignment(horizontal='center', vertical='center')
        total_cell.border = thin_border
        
        # التوقيعات
        signature_row = total_row + 3
        ws[f'A{signature_row}'] = "توقيع الموظف: _____________"
        ws[f'E{signature_row}'] = "توقيع المدير: _____________"
        
        wb.save(filename)
    
    def on_tree_double_click(self, event):
        """فتح الإيصال عند النقر المزدوج"""
        item = self.tree.identify_row(event.y)
        if not item:
            return
        
        index = self.tree.index(item)
        if index < 0 or index >= len(self.expenses):
            return
        
        receipt = self.expenses[index].get('receipt')
        if receipt and os.path.exists(receipt):
            try:
                webbrowser.open(f'file://{os.path.abspath(receipt)}')
            except Exception as e:
                messagebox.showerror("خطأ", f"فشل فتح الإيصال: {e}")
        else:
            messagebox.showinfo("معلومة", "لا يوجد إيصال مرتبط بهذا المصروف.")
    
    def show_profile_window(self):
        """عرض نافذة تعديل البروفايل"""
        win = tk.Toplevel(self.root)
        win.title("تعديل البروفايل")
        win.geometry("550x500")
        win.configure(bg='#16213e')
        win.grab_set()
        
        tk.Label(win, text="تعديل بيانات الحساب", font=('Arial', 16, 'bold'),
                bg='#16213e', fg='#e94560').pack(pady=15)
        
        frame = tk.Frame(win, bg='#16213e')
        frame.pack(pady=10, padx=30, fill='both', expand=True)
        
        # الحقول
        fields_data = [
            ("الاسم الكامل:", 'name', self.current_user.get('name', '')),
            ("رقم الموظف:", 'employee_id', self.current_user.get('employee_id', '')),
            ("اسم الشركة:", 'company_name', self.current_user.get('company_name', '')),
            ("القسم:", 'department', self.current_user.get('department', '')),
            ("البريد الإلكتروني:", 'email', self.current_user.get('email', ''))
        ]
        
        entries = {}
        row = 0
        
        for label, key, value in fields_data:
            tk.Label(frame, text=label, font=('Arial', 10),
                    bg='#16213e', fg='#cbd5e1').grid(row=row, column=0, sticky='e', padx=10, pady=8)
            entry = tk.Entry(frame, font=('Arial', 10), width=35,
                           bg='#0f3460', fg='#ffffff', insertbackground='#ffffff')
            entry.grid(row=row, column=1, padx=10, pady=8)
            entry.insert(0, value)
            entries[key] = entry
            row += 1
        
        # اسم المستخدم (للعرض فقط)
        tk.Label(frame, text="اسم المستخدم:", font=('Arial', 10),
                bg='#16213e', fg='#cbd5e1').grid(row=row, column=0, sticky='e', padx=10, pady=8)
        tk.Label(frame, text=self.current_user.get('username', ''), font=('Arial', 10),
                bg='#16213e', fg='#94a3b8').grid(row=row, column=1, sticky='w', padx=10, pady=8)
        row += 1
        
        # كلمة المرور الجديدة
        tk.Label(frame, text="كلمة المرور الجديدة:", font=('Arial', 10),
                bg='#16213e', fg='#cbd5e1').grid(row=row, column=0, sticky='e', padx=10, pady=8)
        pass_e = tk.Entry(frame, font=('Arial', 10), width=35, show='●',
                         bg='#0f3460', fg='#ffffff', insertbackground='#ffffff')
        pass_e.grid(row=row, column=1, padx=10, pady=8)
        tk.Label(frame, text="(اتركه فارغاً إذا لم ترد التغيير)", font=('Arial', 8),
                bg='#16213e', fg='#64748b').grid(row=row+1, column=1, sticky='w', padx=10)
        row += 2
        
        tk.Label(frame, text="تأكيد كلمة المرور:", font=('Arial', 10),
                bg='#16213e', fg='#cbd5e1').grid(row=row, column=0, sticky='e', padx=10, pady=8)
        pass_conf_e = tk.Entry(frame, font=('Arial', 10), width=35, show='●',
                              bg='#0f3460', fg='#ffffff', insertbackground='#ffffff')
        pass_conf_e.grid(row=row, column=1, padx=10, pady=8)
        row += 1
        
        # وسيلة الدفع الافتراضية
        tk.Label(frame, text="وسيلة الدفع الافتراضية:", font=('Arial', 10),
                bg='#16213e', fg='#cbd5e1').grid(row=row, column=0, sticky='e', padx=10, pady=8)
        pay_cb = ttk.Combobox(frame, font=('Arial', 10), width=32,
                             values=['نقدي', 'فيزا', 'محفظة إلكترونية', 'إنستاباي', 'أخرى'],
                             state='readonly')
        pay_cb.grid(row=row, column=1, padx=10, pady=8)
        pay_cb.set(self.current_user.get('payment_method', 'نقدي'))
        
        def save_profile():
            # التحقق من كلمة المرور
            if pass_e.get().strip():
                if len(pass_e.get().strip()) < 6:
                    messagebox.showerror("خطأ", "كلمة المرور يجب أن تكون 6 أحرف على الأقل!")
                    return
                if pass_e.get() != pass_conf_e.get():
                    messagebox.showerror("خطأ", "كلمة المرور الجديدة غير متطابقة!")
                    return
            
            # التحقق من البريد الإلكتروني
            email = entries['email'].get().strip()
            if email and not self.validate_email(email):
                messagebox.showerror("خطأ", "البريد الإلكتروني غير صحيح!")
                return
            
            # تحديث البيانات
            uname = self.current_user['username']
            self.users_data[uname]['name'] = entries['name'].get().strip()
            self.users_data[uname]['employee_id'] = entries['employee_id'].get().strip()
            self.users_data[uname]['company_name'] = entries['company_name'].get().strip()
            self.users_data[uname]['department'] = entries['department'].get().strip()
            self.users_data[uname]['email'] = email
            self.users_data[uname]['payment_method'] = pay_cb.get()
            
            if pass_e.get().strip():
                self.users_data[uname]['password'] = self.hash_password(pass_e.get().strip())
            
            self.save_users()
            self.current_user = self.users_data[uname].copy()
            self.current_user['username'] = uname
            
            messagebox.showinfo("نجح", "تم تحديث بيانات الحساب بنجاح.")
            win.destroy()
            self.show_main_app()
        
        # أزرار الحفظ والإلغاء
        btn_frame = tk.Frame(win, bg='#16213e')
        btn_frame.pack(pady=15)
        
        tk.Button(btn_frame, text="حفظ التعديلات", font=('Arial', 11, 'bold'),
                 bg='#22c55e', fg='#ffffff', padx=25, pady=8,
                 relief='flat', command=save_profile).pack(side='left', padx=10)
        
        tk.Button(btn_frame, text="إلغاء", font=('Arial', 11),
                 bg='#64748b', fg='#ffffff', padx=25, pady=8,
                 relief='flat', command=win.destroy).pack(side='left', padx=10)
    
    def logout(self):
        """تسجيل الخروج"""
        if messagebox.askyesno("تأكيد", "هل تريد تسجيل الخروج؟"):
            if self.current_user:
                self.save_user_expenses()
            self.current_user = None
            self.expenses = []
            self.show_login_screen()
    
    def on_closing(self):
        """معالجة إغلاق التطبيق"""
        if self.current_user:
            if messagebox.askyesno("تأكيد", "هل تريد حفظ التغييرات والخروج؟"):
                self.save_user_expenses()
                self.root.destroy()
            else:
                self.root.destroy()
        else:
            self.root.destroy()
    
    def run(self):
        """تشغيل التطبيق"""
        self.root.mainloop()


if __name__ == "__main__":
    app = ExpenseTrackerApp()
    app.run()